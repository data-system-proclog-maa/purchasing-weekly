from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os

def style_and_reorder_excel_by_process(path: str):
    """
    apply strings reorder for tidying excel file
    
    Args:
        path (str): location of excel file, no need to input taking from main.py
    """
    print(f"Styling file: {path}")
    try:
        wb = load_workbook(path)
    except FileNotFoundError:
        print(f"File not found: {path}")
        return

    # Define the is_po_file variable based on the path
    is_po_file = 'data_PO_Weekly.xlsx' in os.path.basename(path) 
    
    base_names = [
        'PO_Approved', 'New_RFMfromPO', 'Inprocess_PO',
        'New_RFMfromRFM', 'Inprocess_RFM'
    ]

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    ordered_sheets = []

    if 'Sheet' in wb.sheetnames:
        ordered_sheets.append('Sheet')

    for base in base_names:
        ordered_sheets += [s for s in wb.sheetnames if s.startswith(base) and s not in ordered_sheets]

    for sheet in wb.worksheets:
        title = sheet.title
        
        # --- PR-PO CALCULATION LOGIC ---
        # Only runs if it's the PO file AND the sheet is PO_Approved or a department split
        if is_po_file and title.startswith('PO_Approved'):
            
            # Find the required column indices
            headers = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
            
            try:
                # Find indices (1-based) for the date columns:
                used_approved_date_col_idx = headers.index('used_approved_date') + 1
                po_approval_date_col_idx = headers.index('PO Approval Date') + 1
                
                # Convert indices to Excel column letters
                used_approved_date_col_letter = get_column_letter(used_approved_date_col_idx)
                po_approval_date_col_letter = get_column_letter(po_approval_date_col_idx)
                
                # Determine the position of the new column
                new_col_idx = sheet.max_column + 1
                
                # Set Header in the new column
                sheet.cell(row=1, column=new_col_idx, value='PR-PO')
                sheet.cell(row=1, column=new_col_idx).number_format = 'General'
                
                # Insert formula into every data row (starting from row 2)
                for row_idx in range(2, sheet.max_row + 1):
                    # Formula: = [PO Approval Date] - [used_approved_date]
                    po_cell = f"{po_approval_date_col_letter}{row_idx}"
                    used_approved_cell = f"{used_approved_date_col_letter}{row_idx}"
                    
                    formula = f"={po_cell}-{used_approved_cell}"
                    new_cell = sheet.cell(row=row_idx, column=new_col_idx, value=formula)
                    new_cell.number_format = '0.00'
                
            except ValueError as e:
                print(f"Skipping PR-PO calculation for sheet {title}: Required date column not found.")

        # --- TAB COLORING LOGIC ---
        if title.startswith('PO_Approved'):
            sheet.sheet_properties.tabColor = '00FF00'  # Green
        elif title.startswith('New_RFMfromPO') or title.startswith('New_RFMfromRFM'):
            sheet.sheet_properties.tabColor = '3399FF'  # Blue
        elif title.startswith('Inprocess_PO') or title.startswith('Inprocess_RFM'):
            sheet.sheet_properties.tabColor = 'FFFF00'  # Yellow

        # --- COLUMN HIGHLIGHTING LOGIC ---
        for col_idx, col in enumerate(sheet.iter_cols(min_row=2), start=1):
            header = sheet.cell(row=1, column=col_idx).value
            should_highlight = False

            if title.startswith("PO_Approved") and header == "PO Approval Date":
                should_highlight = True
            elif title.startswith("New_RFMfromPO") and header == "Requisition Approved Date":
                should_highlight = True
            elif title.startswith("Inprocess_PO") and header in ["PO Approval Date", "Requisition Approved Date"]:
                should_highlight = True
            elif title.startswith("New_RFMfromRFM") and header == "Requisition Approved Date":
                should_highlight = True
            elif title.startswith("Inprocess_RFM") and header == "Requisition Approved Date":
                should_highlight = True

            if should_highlight:
                for cell in col:
                    cell.fill = yellow_fill
                    
    # --- REORDERING AND SAVING ---
    wb._sheets = [wb[s] for s in ordered_sheets if s in wb.sheetnames]

    wb.save(path)
    print(f"Styled file saved: {path}")
